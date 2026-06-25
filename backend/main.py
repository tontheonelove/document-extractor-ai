from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import base64, json, os, uuid, csv, io, fitz, math
from datetime import datetime
from filelock import FileLock

# Description: Import Service แบบ Lazy เพื่อป้องกัน Block ตอน Startup
from nuextract_service import nuextract_service

app = FastAPI(title="NuExtract Pro API", version="6.1")

# --- CORS Setup ---
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- Database Config & Locking ---
DB_FILE = "data/db.json"
LOCK_FILE = "data/db.lock"
os.makedirs("data", exist_ok=True)
os.makedirs("data/uploads", exist_ok=True)
os.makedirs("data/icl", exist_ok=True)

def get_db_lock(): 
    """Description: สร้าง FileLock ป้องกัน db.json พังเมื่อเขียนพร้อมกัน"""
    return FileLock(LOCK_FILE, timeout=10)

def load_db():
    """Description: โหลดข้อมูลจาก db.json"""
    if not os.path.exists(DB_FILE): 
        return {"documents": [], "templates": []}
    with open(DB_FILE, "r", encoding="utf-8") as f: 
        return json.load(f)

def save_db(data):
    """Description: บันทึกข้อมูลลง db.json แบบปลอดภัยด้วย FileLock"""
    with get_db_lock():
        with open(DB_FILE, "w", encoding="utf-8") as f: 
            json.dump(data, f, indent=2, ensure_ascii=False)

# --- Pydantic Models ---
class Template(BaseModel):
    id: str
    name: str
    schema: dict
    created_at: str
    icl_examples: List[Dict[str, Any]] = []

class DocumentUpdate(BaseModel):
    result: dict
    status: str
    
class BatchExportRequest(BaseModel):
    doc_ids: List[str]
    format: str = "xlsx"

# --- Helper: PDF to Images ---
def process_pdf_to_base64(file_content: bytes) -> List[str]:
    """Description: แปลง PDF เป็น List ของ Base64 Images (รองรับ Multi-page)"""
    b64_images = []
    try:
        doc = fitz.open(stream=file_content, filetype="pdf")
        for page in doc:
            # dpi=150 เพื่อประหยัด VRAM บน GPU ระดับ Consumer
            pix = page.get_pixmap(dpi=150)
            b64_images.append(base64.b64encode(pix.tobytes("png")).decode("utf-8"))
        return b64_images
    except Exception as e: 
        raise HTTPException(400, f"PDF Processing Error: {str(e)}")

# ==========================================
# STARTUP & SHUTDOWN EVENTS
# ==========================================
@app.on_event("startup")
async def startup_event():
    """Description: แสดงสถานะเมื่อ Server เริ่มทำงาน (ไม่โหลดโมเดลที่นี่เพื่อไม่ให้ Block)"""
    print("=" * 60)
    print("🚀 NuExtract Pro Server v6.1 Started")
    print("📂 Database: data/db.json")
    print("⏳ Model will load on first extraction request (Lazy Loading)")
    print("=" * 60)

@app.on_event("shutdown")
async def shutdown_event():
    """Description: เคลียร์ GPU Memory เมื่อปิด Server"""
    try:
        import torch
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
            print("🧹 GPU memory cleared successfully")
    except ImportError:
        pass

# ==========================================
# ENDPOINTS: TEMPLATES & ICL
# ==========================================
@app.get("/api/templates")
def get_templates(): 
    """Description: ดึงรายการ Template ทั้งหมด"""
    return load_db()["templates"]

@app.post("/api/templates")
def create_template(template: Template):
    """Description: สร้าง Template ใหม่"""
    db = load_db()
    db["templates"].append(template.dict())
    save_db(db)
    return {"status": "success"}

@app.delete("/api/templates/{template_id}")
def delete_template(template_id: str):
    """Description: ลบ Template ตาม ID"""
    db = load_db()
    db["templates"] = [t for t in db["templates"] if t["id"] != template_id]
    save_db(db)
    return {"status": "success"}

@app.post("/api/templates/{template_id}/icl")
async def add_icl_example(
    template_id: str, 
    images: List[UploadFile] = File(...), 
    output_json: str = Form(...)
):
    """Description: เพิ่ม In-Context Learning Example ให้ Template"""
    db = load_db()
    tpl = next((t for t in db["templates"] if t["id"] == template_id), None)
    if not tpl: 
        raise HTTPException(404, "Template not found")

    b64_imgs = []
    for img in images:
        content = await img.read()
        b64_imgs.append(base64.b64encode(content).decode("utf-8"))

    try:
        output_dict = json.loads(output_json)
    except json.JSONDecodeError: 
        raise HTTPException(400, "Invalid JSON output format")

    if "icl_examples" not in tpl: 
        tpl["icl_examples"] = []
    
    tpl["icl_examples"].append({
        "images": b64_imgs, 
        "output": output_dict
    })
    save_db(db)
    return {"status": "success", "count": len(tpl["icl_examples"])}

@app.delete("/api/templates/{template_id}/icl/{icl_idx}")
def delete_icl_example(template_id: str, icl_idx: int):
    """Description: ลบ ICL Example ตาม index"""
    db = load_db()
    tpl = next((t for t in db["templates"] if t["id"] == template_id), None)
    if tpl and 0 <= icl_idx < len(tpl.get("icl_examples", [])):
        tpl["icl_examples"].pop(icl_idx)
        save_db(db)
    return {"status": "success"}

# ==========================================
# ENDPOINTS: EXTRACTION
# ==========================================
@app.post("/api/generate-template")
async def generate_template(description: str = Form(...)):
    """Description: ให้ AI สร้าง JSON Schema จากคำอธิบายภาษาธรรมชาติ"""
    result = nuextract_service.generate_template_from_text(description)
    if result["success"]:
        return {"schema": result["data"]}
    raise HTTPException(400, detail=result["error"])

@app.post("/api/extract")
async def extract_document(
    file: UploadFile = File(...), 
    template_id: Optional[str] = Form(None),
    mode: str = Form("structured"),
    enable_thinking: bool = Form(False),
    instructions: Optional[str] = Form(None)
):
    """Description: Endpoint หลักสำหรับ Extract เอกสาร รองรับ Structured/Markdown/Thinking"""
    db = load_db()
    
    # Validate Template สำหรับ Structured Mode
    template = None
    if mode == "structured":
        if not template_id: 
            raise HTTPException(400, "Template ID required for structured mode")
        template = next((t for t in db["templates"] if t["id"] == template_id), None)
        if not template: 
            raise HTTPException(404, "Template not found")

    # บันทึกไฟล์ลง Disk
    file_id = str(uuid.uuid4())
    content = await file.read()
    file_path = f"data/uploads/{file_id}_{file.filename}"
    with open(file_path, "wb") as f: 
        f.write(content)
    
    # แปลงเป็น Base64 (รองรับ PDF Multi-page)
    if file.filename.lower().endswith(".pdf"):
        b64_images = process_pdf_to_base64(content)
    else:
        b64_images = [base64.b64encode(content).decode("utf-8")]
    
    # ดึง ICL Examples จาก Template
    icl_examples = template.get("icl_examples", []) if template else []
    
    # เรียก AI Service (จะโหลดโมเดลอัตโนมัติถ้ายังไม่โหลด)
    result = nuextract_service.extract(
        images_b64=b64_images, 
        template=template["schema"] if template else None, 
        mode=mode, 
        enable_thinking=enable_thinking,
        instructions=instructions, 
        icl_examples=icl_examples
    )
    
    # บันทึกผลลัพธ์ลง Database
    doc_record = {
        "id": file_id, 
        "name": file.filename,
        "template_name": template["name"] if template else "Markdown-OCR",
        "template_id": template_id,
        "result": result["data"] if result["success"] else {},
        "status": "success" if result["success"] else "failed",
        "mode": mode, 
        "pages": len(b64_images),
        "instructions": instructions,
        "thinking": result.get("thinking"),
        "created_at": datetime.now().isoformat()
    }
    
    db["documents"].insert(0, doc_record)
    save_db(db)
    
    return doc_record

# ==========================================
# ENDPOINTS: DOCUMENTS CRUD & EXPORT
# ==========================================
@app.get("/api/documents")
def get_documents(
    search: Optional[str] = Query(None),
    template_id: Optional[str] = Query(None),  # ✅ เพิ่มตัวกรอง Template
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=10000) 
):
    """
    Description: ดึงรายการเอกสารพร้อม Pagination, Search และ Filter by Template
    """
    docs = load_db()["documents"]
    
    # Search filter (ค้นหาจากชื่อไฟล์)
    if search:
        docs = [d for d in docs if search.lower() in d["name"].lower()]
    
    # ✅ Template filter (กรองตาม Template ID)
    if template_id:
        docs = [d for d in docs if d.get("template_id") == template_id]
    
    # Pagination logic
    total = len(docs)
    total_pages = max(1, math.ceil(total / per_page))
    start_idx = (page - 1) * per_page
    end_idx = start_idx + per_page
    paginated_docs = docs[start_idx:end_idx]
    
    return {
        "documents": paginated_docs,
        "pagination": {
            "page": page,
            "per_page": per_page,
            "total": total,
            "total_pages": total_pages,
            "has_next": page < total_pages,
            "has_prev": page > 1
        }
    }

@app.put("/api/documents/{doc_id}")
def update_document(doc_id: str, update: DocumentUpdate):
    """Description: อัปเดตผลลัพธ์เอกสาร (HITL Editor)"""
    db = load_db()
    for doc in db["documents"]:
        if doc["id"] == doc_id:
            doc["result"] = update.result
            doc["status"] = update.status
            save_db(db)
            return {"status": "success"}
    raise HTTPException(404, "Document not found")

@app.delete("/api/documents/{doc_id}")
def delete_document(doc_id: str):
    """Description: ลบเอกสารตาม ID"""
    db = load_db()
    db["documents"] = [d for d in db["documents"] if d["id"] != doc_id]
    save_db(db)
    return {"status": "success"}

@app.get("/api/documents/{doc_id}/export/{format}")
def export_document(doc_id: str, format: str):
    """Description: Export เอกสารเป็น JSON, CSV, Markdown หรือ Excel (XLSX)"""
    db = load_db()
    doc = next((d for d in db["documents"] if d["id"] == doc_id), None)
    if not doc: 
        raise HTTPException(404, "Document not found")
    
    # JSON Export
    if format == "json":
        return JSONResponse(
            content=doc["result"], 
            headers={"Content-Disposition": f"attachment; filename={doc['name']}.json"}
        )
    
    # Markdown Export
    elif format == "md":
        md_content = doc["result"] if isinstance(doc["result"], str) else json.dumps(doc["result"], indent=2)
        return StreamingResponse(
            io.BytesIO(md_content.encode("utf-8")), 
            media_type="text/markdown", 
            headers={"Content-Disposition": f"attachment; filename={doc['name']}.md"}
        )
        
    # CSV Export (Flattened)
    elif format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        
        def flatten(data, prefix=""):
            items = []
            if isinstance(data, dict):
                for k, v in data.items(): 
                    items.extend(flatten(v, f"{prefix}{k}."))
            elif isinstance(data, list):
                for i, item in enumerate(data): 
                    items.extend(flatten(item, f"{prefix}[{i}]."))
            else: 
                items.append((prefix.rstrip("."), str(data) if data is not None else ""))
            return items
        
        flat_data = flatten(doc["result"])
        if flat_data:
            writer.writerow([k for k, v in flat_data])
            writer.writerow([v for k, v in flat_data])
        else:
            writer.writerow(["empty"])
            writer.writerow(["no data"])
        
        # ✅ เพิ่ม UTF-8 BOM เพื่อให้ Excel อ่านภาษาไทยได้ถูกต้อง
        csv_content = '\ufeff' + output.getvalue()
        output_bom = io.BytesIO(csv_content.encode("utf-8"))
        
        return StreamingResponse(
            output_bom,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={doc['name']}.csv"}
        )
        
    # Excel Export (.xlsx)
    elif format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"
        
        def flatten(data, prefix=""):
            items = []
            if isinstance(data, dict):
                for k, v in data.items(): 
                    items.extend(flatten(v, f"{prefix}{k}."))
            elif isinstance(data, list):
                for i, item in enumerate(data): 
                    items.extend(flatten(item, f"{prefix}[{i}]."))
            else: 
                items.append((prefix.rstrip("."), str(data) if data is not None else ""))
            return items
        
        flat_data = flatten(doc["result"])
        
        if flat_data:
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
            
            for col_idx, (key, value) in enumerate(flat_data, start=1):
                header_cell = ws.cell(row=1, column=col_idx, value=key)
                header_cell.font = header_font
                header_cell.fill = header_fill
                header_cell.alignment = Alignment(horizontal="center")
                
                data_cell = ws.cell(row=2, column=col_idx, value=value)
                data_cell.alignment = Alignment(horizontal="left")
            
            # Auto-adjust column width
            for col_idx, (key, value) in enumerate(flat_data, start=1):
                max_len = max(len(str(key)), len(str(value)))
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        else:
            ws.cell(row=1, column=1, value="No data")
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return StreamingResponse(
            excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={doc['name']}.xlsx"}
        )
    
    raise HTTPException(400, "Invalid format. Use: json, csv, md, or xlsx")


@app.post("/api/documents/export-batch")
def export_batch_documents(request: BatchExportRequest):
    """
    Description: Export หลายเอกสารเป็นไฟล์เดียว (Batch Export)
    """
    doc_ids = request.doc_ids
    format = request.format
    
    if not doc_ids:
        raise HTTPException(400, "doc_ids is required")
    
    db = load_db()
    
    # ดึงเอกสารทั้งหมดที่เลือก (รักษา order ตาม doc_ids)
    docs = []
    for doc_id in doc_ids:
        doc = next((d for d in db["documents"] if d["id"] == doc_id), None)
        if doc:
            docs.append(doc)
    
    if not docs:
        raise HTTPException(404, "No documents found")
    
    # Flatten ข้อมูลทุก doc เป็น rows
    rows = []
    all_keys = []
    
    for doc in docs:
        flat_data = flatten_dict(doc["result"])
        row = {k: v for k, v in flat_data}
        rows.append(row)
        
        for k in row.keys():
            if k not in all_keys:
                all_keys.append(k)
    
    if not all_keys:
        all_keys = ["empty"]
        for row in rows:
            row["empty"] = ""
    
    filename_base = f"batch_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    # ===== CSV Export =====
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(all_keys)
        for row in rows:
            writer.writerow([row.get(k, "") for k in all_keys])
        
        csv_content = '\ufeff' + output.getvalue()
        output_bom = io.BytesIO(csv_content.encode("utf-8"))
        
        return StreamingResponse(
            output_bom,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.csv"}
        )
    
    # ===== Excel Export (.xlsx) =====
    elif format == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Batch Export"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        for col_idx, key in enumerate(all_keys, start=1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(all_keys, start=1):
                value = row.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        
        for col_idx, key in enumerate(all_keys, start=1):
            max_len = len(str(key))
            for row in rows:
                val_len = len(str(row.get(key, "")))
                if val_len > max_len:
                    max_len = val_len
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50)
        
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(all_keys)).column_letter}1"
        
        excel_bytes = io.BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        
        return StreamingResponse(
            excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename_base}.xlsx"}
        )
    
    raise HTTPException(400, "Invalid format. Use: csv or xlsx")


# Helper function สำหรับ flatten nested dict
def flatten_dict(data, prefix=""):
    """Description: Flatten nested dict/list เป็น key-value pairs"""
    items = []
    if isinstance(data, dict):
        for k, v in data.items():
            items.extend(flatten_dict(v, f"{prefix}{k}."))
    elif isinstance(data, list):
        for i, item in enumerate(data):
            items.extend(flatten_dict(item, f"{prefix}[{i}]."))
    else:
        items.append((prefix.rstrip("."), str(data) if data is not None else ""))
    return items

# ==========================================
# HEALTH CHECK ENDPOINT
# ==========================================
@app.get("/api/health")
def health_check():
    """
    Description: Health check endpoint สำหรับ Frontend status indicator
    Return: status + GPU info (ถ้ามี)
    """
    try:
        import torch
        gpu_info = None
        if torch.cuda.is_available():
            gpu_info = {
                "name": torch.cuda.get_device_name(0),
                "vram_total_gb": round(torch.cuda.get_device_properties(0).total_mem / 1e9, 1),
                "vram_used_gb": round(torch.cuda.memory_allocated(0) / 1e9, 2)
            }
        
        return {
            "status": "online",
            "version": "6.1",
            "model": "numind/NuExtract3",
            "gpu": gpu_info
        }
    except Exception as e:
        return {
            "status": "online",
            "version": "6.1",
            "model": "numind/NuExtract3",
            "gpu": None,
            "note": f"GPU info unavailable: {str(e)}"
        }

@app.get("/")
def root():
    """Description: Root endpoint"""
    return {
        "message": "NuExtract Pro API is running", 
        "version": "6.1",
        "endpoints": {
            "health": "/api/health",
            "documents": "/api/documents",
            "templates": "/api/templates",
            "extract": "/api/extract"
        }
    }